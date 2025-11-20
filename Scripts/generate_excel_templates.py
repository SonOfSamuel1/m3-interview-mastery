#!/usr/bin/env python3
"""
Generate Excel templates for M1 module:
1. Company Evaluation Scorecard
2. Target Company List Template (with 50+ pre-populated companies)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_company_evaluation_scorecard():
    """Create Company Evaluation Scorecard Excel template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Scorecard"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    criteria_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = 'COMPANY EVALUATION SCORECARD'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')

    # Instructions
    ws['A2'] = 'Instructions: Rate each company 1-5 for each criterion (1=Poor, 5=Excellent). Weighted score calculates automatically.'
    ws.merge_cells('A2:H2')
    ws['A2'].font = Font(size=10, italic=True)

    # Headers (Row 4)
    headers = ['Criterion', 'Weight', 'Company 1', 'Company 2', 'Company 3', 'Company 4', 'Company 5']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Criteria and weights
    criteria = [
        ('Compensation (OTE + Equity)', 0.25),
        ('Company Stage / Risk Level', 0.20),
        ('Culture Fit', 0.15),
        ('Product / Market Position', 0.15),
        ('Career Growth Opportunity', 0.10),
        ('Work-Life Balance', 0.10),
        ('Location / Remote Policy', 0.05)
    ]

    # Populate criteria
    row = 5
    for criterion, weight in criteria:
        ws.cell(row=row, column=1).value = criterion
        ws.cell(row=row, column=1).fill = criteria_fill
        ws.cell(row=row, column=1).border = border

        ws.cell(row=row, column=2).value = weight
        ws.cell(row=row, column=2).number_format = '0%'
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).border = border

        # Empty score cells for companies 1-5
        for col in range(3, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row += 1

    # Weighted Score row
    ws.cell(row=row, column=1).value = 'WEIGHTED SCORE'
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws.cell(row=row, column=1).border = border

    ws.cell(row=row, column=2).value = '100%'
    ws.cell(row=row, column=2).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws.cell(row=row, column=2).border = border
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')

    # Formulas for weighted scores
    for col in range(3, 8):
        formula = f'=SUMPRODUCT(B5:B11,{get_column_letter(col)}5:{get_column_letter(col)}11)'
        cell = ws.cell(row=row, column=col)
        cell.value = formula
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = '0.00'

    # Company names row (for reference)
    ws.cell(row=3, column=1).value = 'Company Name:'
    ws.cell(row=3, column=1).font = Font(bold=True)
    for col in range(3, 8):
        cell = ws.cell(row=3, column=col)
        cell.value = f'Enter Name'
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    for col in range(3, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save
    output_path = "/Users/terrancebrandon/Desktop/Active Offer/AO- Course Content/Active Offer- Course Material/External Assets/Company-Evaluation-Scorecard.xlsx"
    wb.save(output_path)
    print(f"✓ Saved Company Evaluation Scorecard: {output_path}")

def create_target_company_list():
    """Create Target Company List Template with 50+ pre-populated companies"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Target Company List"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = 'TARGET COMPANY LIST TEMPLATE'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:I1')

    # Instructions
    ws['A2'] = 'Instructions: Review pre-populated companies below. Add your own companies. Use Priority column to rank (A/B/C). Track your application progress.'
    ws.merge_cells('A2:I2')
    ws['A2'].font = Font(size=10, italic=True)

    # Headers
    headers = ['Company Name', 'Profile Type', 'Revenue Tier', 'Est. OTE Range', 'Headquarters', 'Remote Policy', 'Priority (A/B/C)', 'Status', 'Notes']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Pre-populated companies data
    companies = [
        # Enterprise Leaders
        ('Salesforce', 'Enterprise Leaders', '$10B+', '$300k-$600k', 'San Francisco, CA', 'Hybrid', 'A', '', ''),
        ('Microsoft (Dynamics)', 'Enterprise Leaders', '$10B+', '$250k-$500k', 'Redmond, WA', 'Flexible', 'A', '', ''),
        ('Oracle', 'Enterprise Leaders', '$10B+', '$280k-$550k', 'Austin, TX', 'Hybrid', 'B', '', ''),
        ('Google Cloud', 'Enterprise Leaders', '$10B+', '$280k-$550k', 'Sunnyvale, CA', 'Hybrid', 'A', '', ''),
        ('AWS', 'Enterprise Leaders', '$10B+', '$300k-$650k', 'Seattle, WA', 'Flexible', 'A', '', ''),
        ('Adobe', 'Enterprise Leaders', '$10B+', '$250k-$450k', 'San Jose, CA', 'Hybrid', 'B', '', ''),
        ('SAP', 'Enterprise Leaders', '$10B+', '$280k-$500k', 'Newtown Square, PA', 'Hybrid', 'B', '', ''),
        ('ServiceNow', 'Enterprise Leaders', '$5B+', '$300k-$550k', 'Santa Clara, CA', 'Hybrid', 'A', '', ''),
        ('Workday', 'Enterprise Leaders', '$5B+', '$280k-$500k', 'Pleasanton, CA', 'Hybrid', 'B', '', ''),
        ('HubSpot', 'Enterprise Leaders', '$1B+', '$200k-$400k', 'Cambridge, MA', 'Remote-First', 'A', '', ''),
        ('Zoom', 'Enterprise Leaders', '$4B+', '$220k-$450k', 'San Jose, CA', 'Hybrid', 'B', '', ''),
        ('Atlassian', 'Enterprise Leaders', '$2B+', '$250k-$450k', 'San Francisco, CA', 'Distributed', 'B', '', ''),
        ('Snowflake', 'Enterprise Leaders', '$2B+', '$300k-$600k', 'Bozeman, MT', 'Flexible', 'A', '', ''),
        ('Datadog', 'Enterprise Leaders', '$1.5B+', '$280k-$500k', 'New York, NY', 'Hybrid', 'B', '', ''),
        ('Splunk', 'Enterprise Leaders', '$3B+', '$270k-$500k', 'San Francisco, CA', 'Flexible', 'B', '', ''),

        # Growth Champions
        ('Gong', 'Growth Champions', '$200M+', '$250k-$400k', 'San Francisco, CA', 'Hybrid', 'A', '', ''),
        ('Notion', 'Growth Champions', '$100M+', '$200k-$380k', 'San Francisco, CA', 'Remote-First', 'A', '', ''),
        ('Airtable', 'Growth Champions', '$150M+', '$220k-$400k', 'San Francisco, CA', 'Hybrid', 'A', '', ''),
        ('Miro', 'Growth Champions', '$200M+', '$240k-$420k', 'San Francisco, CA', 'Distributed', 'A', '', ''),
        ('Webflow', 'Growth Champions', '$100M+', '$200k-$350k', 'San Francisco, CA', 'Remote-First', 'B', '', ''),
        ('Rippling', 'Growth Champions', '$200M+', '$230k-$400k', 'San Francisco, CA', 'Hybrid', 'A', '', ''),
        ('DataRobot', 'Growth Champions', '$300M+', '$260k-$450k', 'Boston, MA', 'Hybrid', 'B', '', ''),
        ('Mixpanel', 'Growth Champions', '$100M+', '$210k-$380k', 'San Francisco, CA', 'Remote-First', 'B', '', ''),
        ('Superhuman', 'Growth Champions', '$30M+', '$180k-$320k', 'San Francisco, CA', 'Remote-First', 'B', '', ''),
        ('Figma', 'Growth Champions', '$400M+', '$280k-$480k', 'San Francisco, CA', 'Distributed', 'A', '', ''),
        ('Carta', 'Growth Champions', '$150M+', '$220k-$400k', 'San Francisco, CA', 'Hybrid', 'B', '', ''),
        ('Amplitude', 'Growth Champions', '$200M+', '$240k-$420k', 'San Francisco, CA', 'Remote-First', 'B', '', ''),
        ('Braze', 'Growth Champions', '$300M+', '$250k-$430k', 'New York, NY', 'Hybrid', 'B', '', ''),
        ('UiPath', 'Growth Champions', '$1B+', '$280k-$500k', 'New York, NY', 'Hybrid', 'A', '', ''),
        ('Toast', 'Growth Champions', '$3B+', '$250k-$450k', 'Boston, MA', 'Hybrid', 'B', '', ''),

        # Venture-Backed Startups
        ('Ramp', 'Venture-Backed Startups', '$50M+', '$160k-$240k', 'New York, NY', 'Hybrid', 'C', '', ''),
        ('Brex', 'Venture-Backed Startups', '$200M+', '$180k-$260k', 'San Francisco, CA', 'Hybrid', 'B', '', ''),
        ('Deel', 'Venture-Backed Startups', '$300M+', '$170k-$250k', 'San Francisco, CA', 'Remote-First', 'B', '', ''),
        ('Mercury', 'Venture-Backed Startups', '$50M+', '$150k-$230k', 'San Francisco, CA', 'Remote-First', 'C', '', ''),
        ('Hex', 'Venture-Backed Startups', '<$50M', '$140k-$220k', 'San Francisco, CA', 'Remote-First', 'C', '', ''),
        ('Clay', 'Venture-Backed Startups', '<$50M', '$145k-$220k', 'New York, NY', 'Remote-First', 'C', '', ''),
        ('Watershed', 'Venture-Backed Startups', '<$50M', '$150k-$230k', 'San Francisco, CA', 'Remote-First', 'C', '', ''),
        ('Crossbeam', 'Venture-Backed Startups', '<$50M', '$140k-$210k', 'Philadelphia, PA', 'Remote-First', 'C', '', ''),
        ('Modal', 'Venture-Backed Startups', '<$50M', '$140k-$220k', 'San Francisco, CA', 'Remote-First', 'C', '', ''),
        ('Juni Learning', 'Venture-Backed Startups', '<$50M', '$135k-$200k', 'San Francisco, CA', 'Remote-First', 'C', '', ''),
        ('Stytch', 'Venture-Backed Startups', '<$50M', '$145k-$225k', 'San Francisco, CA', 'Remote-First', 'C', '', ''),
        ('Vanta', 'Venture-Backed Startups', '$50M+', '$155k-$235k', 'San Francisco, CA', 'Remote-First', 'B', '', ''),
        ('Pilot', 'Venture-Backed Startups', '<$50M', '$140k-$215k', 'San Francisco, CA', 'Remote-First', 'C', '', ''),
        ('Sendbird', 'Venture-Backed Startups', '$50M+', '$150k-$230k', 'San Mateo, CA', 'Hybrid', 'C', '', ''),
        ('Loom', 'Venture-Backed Startups', '$100M+', '$165k-$250k', 'San Francisco, CA', 'Remote-First', 'B', '', ''),

        # More Enterprise Leaders
        ('VMware', 'Enterprise Leaders', '$10B+', '$270k-$500k', 'Palo Alto, CA', 'Hybrid', 'B', '', ''),
        ('Cisco (Collaboration)', 'Enterprise Leaders', '$10B+', '$260k-$480k', 'San Jose, CA', 'Hybrid', 'B', '', ''),
        ('IBM (Cloud)', 'Enterprise Leaders', '$10B+', '$250k-$450k', 'Armonk, NY', 'Flexible', 'C', '', ''),
        ('Box', 'Enterprise Leaders', '$1B+', '$230k-$420k', 'Redwood City, CA', 'Hybrid', 'B', '', ''),
        ('Dropbox', 'Enterprise Leaders', '$2B+', '$240k-$440k', 'San Francisco, CA', 'Virtual First', 'B', '', ''),
    ]

    # Populate company data
    row = 5
    for company_data in companies:
        for col, value in enumerate(company_data, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            if col in [1, 2, 3, 5, 6]:  # Text columns
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif col == 4:  # OTE Range
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 7:  # Priority
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if value == 'A':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value == 'B':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif value == 'C':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1

    # Add some blank rows for user to add companies
    for i in range(10):
        for col in range(1, 10):
            cell = ws.cell(row=row + i, column=col)
            cell.border = border

    # Column widths
    ws.column_dimensions['A'].width = 25  # Company Name
    ws.column_dimensions['B'].width = 22  # Profile Type
    ws.column_dimensions['C'].width = 15  # Revenue Tier
    ws.column_dimensions['D'].width = 15  # OTE Range
    ws.column_dimensions['E'].width = 20  # Headquarters
    ws.column_dimensions['F'].width = 15  # Remote Policy
    ws.column_dimensions['G'].width = 13  # Priority
    ws.column_dimensions['H'].width = 15  # Status
    ws.column_dimensions['I'].width = 30  # Notes

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Save
    output_path = "/Users/terrancebrandon/Desktop/Active Offer/AO- Course Content/Active Offer- Course Material/External Assets/Target-Company-List-Template.xlsx"
    wb.save(output_path)
    print(f"✓ Saved Target Company List Template: {output_path}")

def main():
    print("=" * 60)
    print("Generating Excel Templates")
    print("=" * 60)
    print()

    # Ensure External Assets directory exists
    import os
    os.makedirs("/Users/terrancebrandon/Desktop/Active Offer/AO- Course Content/Active Offer- Course Material/External Assets", exist_ok=True)

    create_company_evaluation_scorecard()
    create_target_company_list()

    print()
    print("✓ All Excel templates generated successfully!")

if __name__ == "__main__":
    main()
